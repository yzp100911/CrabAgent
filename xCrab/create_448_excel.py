import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os

# 创建工作簿
wb = openpyxl.Workbook()

# 第一个工作表：填充后的模板
ws1 = wb.active
ws1.title = "退租申请表"

# 设置列宽
ws1.column_dimensions['A'].width = 20
ws1.column_dimensions['B'].width = 40
ws1.column_dimensions['C'].width = 15
ws1.column_dimensions['D'].width = 15

# 标题样式
title_font = Font(name='微软雅黑', size=14, bold=True)
header_font = Font(name='微软雅黑', size=11, bold=True)
normal_font = Font(name='微软雅黑', size=10)

# 填充数据
data = {
    "店铺名称": "南城店",
    "房间号": "448",
    "租客姓名": "陈丹",
    "租客电话": "18419236438",
    "签约时间": "2026年04月12日",
    "签约房号": "448",
    "房屋押金": 550.00,
    "房卡押金": 0.00,
    "租期截止日期": "2026年06月10日",
    "搬离日期": "2026年06月04日",
    "冷水用量(吨)": 2.00,
    "冷水单价(元/吨)": 4.50,
    "冷水费(元)": 9.00,
    "二次加压用量(吨)": 2.00,
    "二次加压单价(元/吨)": 1.50,
    "二次加压费(元)": 3.00,
    "热水用量(吨)": 1.00,
    "热水单价(元/吨)": 6.00,
    "热水费(元)": 6.00,
    "退房卫生费(元)": 0.00,
    "物品赔偿费(元)": 0.00,
    "其他费用(元)": 73.33,
    "其他费用备注": "房租",
    "总扣除费用(元)": 91.33,
    "应退押金(元)": 550.00,
    "实际退款(元)": 458.67,
    "退租原因": "特殊退租",
    "退租备注": "南城店6月15号之后会搬迁解散，故不用管是否到期。免收退房卫生费。"
}

# 写入数据
row = 1
ws1.cell(row=row, column=1, value="有米公寓退租申请表").font = title_font
ws1.merge_cells('A1:D1')
row = 3

for key, value in data.items():
    ws1.cell(row=row, column=1, value=key).font = header_font
    ws1.cell(row=row, column=2, value=value).font = normal_font
    row += 1

# 添加边框
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in ws1.iter_rows(min_row=1, max_row=row-1, min_col=1, max_col=2):
    for cell in row:
        cell.border = thin_border

# 第二个工作表：费用明细
ws2 = wb.create_sheet("费用明细")
ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 15
ws2.column_dimensions['D'].width = 15
ws2.column_dimensions['E'].width = 15

# 费用明细数据
expenses = [
    ["费用项目", "用量", "单价", "金额", "备注"],
    ["冷水费", 2.00, 4.50, 9.00, ""],
    ["二次加压和维护损耗", 2.00, 1.50, 3.00, ""],
    ["热水费", 1.00, 6.00, 6.00, ""],
    ["退房卫生费", "", "", 0.00, "免收"],
    ["物品赔偿费", "", "", 0.00, ""],
    ["房租", "", "", 73.33, "租金"],
    ["总计", "", "", 91.33, ""]
]

for i, row_data in enumerate(expenses, 1):
    for j, value in enumerate(row_data, 1):
        cell = ws2.cell(row=i, column=j, value=value)
        if i == 1:
            cell.font = header_font
        else:
            cell.font = normal_font
        cell.border = thin_border

# 保存文件
output_path = "448房退租申请表.xlsx"
wb.save(output_path)
print(f"Excel文件已保存: {output_path}")