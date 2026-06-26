import openpyxl
from openpyxl.utils import get_column_letter

filepath = '/www/wwwroot/eclaw/uploads/退租申请表_443房.xlsx'
wb = openpyxl.load_workbook(filepath)
sheet = wb.active

# 字段值映射（基于系统实际数据）
data = {
    'B4':  '443',                                  # 房间号
    'B5':  '杨锟明',                               # 租客姓名
    'B6':  '2025/02/17 - 2026/05/31',              # 租期起止日
    'B7':  '2026/06/14',                           # 退房时间
    'B8':  850,                                    # 房屋押金
    'B9':  0,                                      # 门卡押金
    'B10': 100,                                    # 家电押金
    'B11': 0,                                      # 剩余房屋租金
    'B12': '',                                     # 剩余房屋租金备注
    'B13': 0,                                      # 其他应退还金额
    'B14': '',                                     # 其他应退还金额备注
    'B15': 396.67,                                 # 房屋租金（欠费）
    'B16': '2026.6.1-2026.6.14 房租14天×28.33',  # 房屋租金备注
    'B17': 1.0,                                    # 冷水用量(吨)
    'B19': 4.5,                                    # 冷水费(元) 直接填值
    'B20': 1.0,                                    # 二次加压和维护损耗用量(吨)
    'B22': 1.5,                                    # 二次加压和维护损耗费(元)
    'B23': 0.80,                                   # 热水用量(吨)
    'B25': 4.80,                                   # 热水费(元)
    'B26': 0,                                      # 退房卫生费
    'B27': 0,                                      # 物品赔偿费
    'B28': 0,                                      # 其他费用
    'B29': '',                                     # 其他费用备注
}

for cell, val in data.items():
    sheet[cell] = val

# 在备注行（C16/D16/E16区域）追加信息 - 在已存在的B33区域追加补充
# 在B33 后追加详细备注
remark = (
    '\n退房时间：2026/06/14；'
    '退租原因：特殊退租；'
    '租客电话：13535305604；'
    '房东：卢健民/13342678002；'
    '经办人：庞世雄；'
    '交易时间：2026-06-17 14:39:04（银行卡转账，经办人卢健民）。'
    '押金总额950元（房屋押850+家电押100），退房结算实际退款542.53元。'
    '房屋租金补收396.67元（850÷30×14天，2026.6.1-6.14）；'
    '冷水1吨4.5元（423→424）；二次加压1吨1.5元；热水0.8吨4.8元（153→153.8）；'
    '公共用电0度0元。总扣除407.47元，应退950元，实际退款542.53元。'
)
if sheet['B33'].value:
    sheet['B33'] = str(sheet['B33'].value) + remark
else:
    sheet['B33'] = remark

wb.save(filepath)
print(f"已保存到 {filepath}")

# 重新读取验证
wb2 = openpyxl.load_workbook(filepath)
sh = wb2.active
print("\n=== 验证写入结果 ===")
for r in range(1, sh.max_row + 1):
    for c in range(1, sh.max_column + 1):
        v = sh.cell(row=r, column=c).value
        if v is not None and v != '':
            print(f"{get_column_letter(c)}{r}: {v}")
wb2.close()
